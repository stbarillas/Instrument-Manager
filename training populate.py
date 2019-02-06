import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')

import django
django.setup()
from blog.models import TrainingDocument, JobTitle, Profile
from django.contrib.auth.models import User

import openpyxl

def populatetraining():
    wb = openpyxl.load_workbook('example.xlsm')
    sheet = wb.get_sheet_by_name('Sheet1')
    doc_name = []
    frequency = []
    for title in sheet['B6':'B45']:
        for cell in title:
            doc_name.append(cell.value)

    for renewal in sheet['D6':'D45']:
        for cell in renewal:
            frequency.append(cell.value)

    count = 0
    while count < len(doc_name):
        add_training(doc_name[count],frequency[count])
        count += 1

def populatejobs():
    wb = openpyxl.load_workbook('example.xlsm')
    sheet = wb.get_sheet_by_name('Sheet1')

    doc_name = []
    for title in sheet['B6':'B45']:
        for cell in title:
            doc_name.append(cell.value)

    list = []
    for col in sheet.iter_cols(min_row=5, min_col=7, max_col=26, max_row=45):
        column = []
        for cell in col:
            column.append(cell.value)
        list.append(column)

    for column in list:
        count = 1
        job = JobTitle.objects.get_or_create(title=column[0])[0]
        required_training = []
        while count < 38:
            # print(column[count])
            if column[count] == 'R':
                doc_id = TrainingDocument.objects.get(document_name = doc_name[count - 1]).id
                required_training.append(doc_id)
            count += 1
        job.required_training = required_training
        job.save()

def populateusers():
    wb = openpyxl.load_workbook('example.xlsm')
    sheet = wb.get_sheet_by_name('Sheet1')

    users = []
    for user in sheet['AB5':'IO5']:
        for cell in user:
            users.append(cell.value)

    job_titles = []
    for title in sheet['AB3':'IO3']:
        for cell in title:
            job_titles.append(cell.value)

    manager = []
    for man in sheet['AB4':'IO4']:
        for cell in man:
            manager.append(cell.value)

    count = 0
    for user in users:
        user_creation = User.objects.get_or_create(username=user)[0]
        split = user.split()
        user_intance = User.objects.get(username=user)
        user_creation.save()
        profile_creation = Profile.objects.get_or_create(user = user_intance)[0]
        profile_creation.first_name = split[0]
        profile_creation.last_name = split[1]
        print(user)
        print(job_titles[count])
        print(manager[count])
        profile_creation.job_title = JobTitle.objects.get(title= job_titles[count])
        profile_creation.manager = User.objects.get(username=manager[count])
        count += 1
        profile_creation.save()



def add_training(name, renewal):
    training = TrainingDocument.objects.get_or_create(document_name=name)[0]
    training.document_name = str(name)
    training.document_renewal_years= str(renewal)
    training.document_copy = 'images/singlequad.jpg'
    training.save()
    return

# Start execution here!
if __name__ == '__main__':
    print("Starting Training population script...")
    populatetraining()
    print("Starting Job Title population script...")
    populatejobs()
    print("Starting User population script...")
    populateusers()
